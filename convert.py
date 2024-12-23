import logging
from pathlib import Path
from typing import Tuple, Optional, List
from docx import Document
from openpyxl import Workbook
import re
from docx.text.run import Run
from docx.oxml.text.run import CT_R

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('converter.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class TextProcessor:
    @staticmethod
    def is_roman_numeral(text: str) -> bool:
        """Проверяет, является ли текст римской цифрой."""
        roman_pattern = r'^(I|II|III|IV|V|VI|VII|VIII|IX|X)$'
        return bool(re.match(roman_pattern, text.strip()))

    @staticmethod
    def normalize_text(text: str) -> str:
        """Нормализует текст, удаляя лишние пробелы."""
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def clean_first_word(word: str) -> str:
        """Очищает первое слово от знаков препинания."""
        return re.sub(r'[,:]+$', '', word).strip()

class FormatProcessor:
    @staticmethod
    def get_run_formatting(run: Run) -> Optional[str]:
        """Определяет форматирование текстового фрагмента."""
        if run.bold and run.italic:
            return 'bold italic'
        elif run.bold:
            return 'bold'
        elif run.italic:
            return 'italic'
        return None

    def process_runs(self, paragraph, skip_words: int = 0) -> str:
        """Обрабатывает форматирование текстовых фрагментов параграфа."""
        formatted_text = []
        current_format = None
        current_text = []
        words_processed = 0
        
        for run in paragraph.runs:
            run_text = run.text
            if not run_text.strip():
                if current_text and run_text.isspace():
                    current_text.append(run_text)
                continue
                
            if words_processed < skip_words:
                words = run_text.split()
                if len(words) + words_processed <= skip_words:
                    words_processed += len(words)
                    continue
                else:
                    remaining_words = skip_words - words_processed
                    run_text = ' '.join(words[remaining_words:])
                    words_processed = skip_words
            
            run_format = self.get_run_formatting(run)
            
            if run_format != current_format:
                if current_text:
                    text = ''.join(current_text)
                    if current_format:
                        text = f'<span class="{current_format}">{text}</span>'
                    formatted_text.append(text)
                    current_text = []
                current_format = run_format
            
            # Сохраняем оригинальный текст без модификации пробелов
            current_text.append(run_text)
        
        if current_text:
            text = ''.join(current_text)
            if current_format:
                text = f'<span class="{current_format}">{text}</span>'
            formatted_text.append(text)
        
        # Объединяем текст, сохраняя оригинальные пробелы
        return ''.join(formatted_text)

class DocumentProcessor:
    def __init__(self):
        self.text_processor = TextProcessor()
        self.format_processor = FormatProcessor()

    def process_paragraph(self, paragraph) -> Tuple[str, str]:
        """Обрабатывает параграф документа."""
        text = self.text_processor.normalize_text(paragraph.text)
        words = text.split()
        if not words:
            return "", ""
        
        first_word = self.text_processor.clean_first_word(words[0].upper())
        skip_words = 1
        has_punctuation = None
        
        if len(words) > 1:
            second_word = words[1]
            second_word_clean = self.text_processor.clean_first_word(second_word)
            
            if self.text_processor.is_roman_numeral(second_word_clean):
                first_word = f"{first_word} {second_word_clean}"
                skip_words = 2
                
                if second_word.endswith(':'):
                    has_punctuation = ':'
                elif second_word.endswith(','):
                    has_punctuation = ','
                elif len(words) > 2 and (words[2] == ':' or words[2] == ','):
                    has_punctuation = words[2]
                    skip_words = 3
        else:
            if words[0].endswith(':'):
                has_punctuation = ':'
            elif words[0].endswith(','):
                has_punctuation = ','
        
        formatted_content = self.format_processor.process_runs(paragraph, skip_words)
        formatted_content = re.sub(r'\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b[:, ]?', '', formatted_content)
        formatted_content = re.sub(r'<(?!/?span)[^>]+>', '', formatted_content)
        
        if has_punctuation:
            if '<span' in formatted_content:
                formatted_content = re.sub(r'(<span[^>]*>)(.*?)(</span>)', 
                                       rf'\1\2{has_punctuation}\3', 
                                       formatted_content, 
                                       count=1)
            else:
                formatted_content = f'{has_punctuation} {formatted_content}'
        
        formatted_content = re.sub(r'\s+', ' ', formatted_content).strip()
        formatted_content = re.sub(r':{2,}', ':', formatted_content)
        formatted_content = re.sub(r',{2,}', ',', formatted_content)
        
        return first_word, formatted_content

class BatchConverter:
    def __init__(self):
        self.document_processor = DocumentProcessor()
        self.input_dir = Path('input')
        self.output_dir = Path('output')
        
    def setup_directories(self) -> None:
        """Создает необходимые директории, если они не существуют."""
        self.input_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)
        
    def get_word_files(self) -> List[Path]:
        """Получает список всех Word файлов в директории input."""
        return list(self.input_dir.glob('*.docx'))
        
    def convert_single_file(self, doc_path: Path) -> None:
        """Конвертирует один Word файл в Excel."""
        try:
            output_path = self.output_dir / f"{doc_path.stem}.xlsx"
            logger.info(f"Начало конвертации: {doc_path.name}")
            
            doc = Document(doc_path)
            wb = Workbook()
            ws = wb.active
            
            # Добавляем заголовки
            headers = ['articleid', 'articlecat', 'articletitle', 'articleintrotext']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            row = 2
            article_id = 1  # Начальный номер для articleid
            article_cat = doc_path.stem.upper()  # Берем первую букву имени файла в верхнем регистре
            
            for para in doc.paragraphs:
                text = self.document_processor.text_processor.normalize_text(para.text)
                if text:
                    title, content = self.document_processor.process_paragraph(para)
                    if title:
                        # articleid (порядковый номер)
                        ws.cell(row=row, column=1, value=article_id)
                        # articlecat (буква файла)
                        ws.cell(row=row, column=2, value=article_cat)
                        # articletitle
                        ws.cell(row=row, column=3, value=title)
                        # articleintrotext
                        ws.cell(row=row, column=4, value=f"<p>{content}</p>")
                        
                        article_id += 1  # Увеличиваем порядковый номер
                        row += 1
            
            # Автоматическая настройка ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)  # Ограничиваем максимальную ширину
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            logger.info(f"Файл {doc_path.name} успешно обработан. Параграфов: {row - 2}")
            
        except Exception as e:
            logger.error(f"Ошибка при обработке файла {doc_path.name}: {str(e)}", exc_info=True)
            raise
            
    def process_all_files(self) -> None:
        """Обрабатывает все Word файлы в директории input."""
        try:
            self.setup_directories()
            word_files = self.get_word_files()
            
            if not word_files:
                logger.warning("Word файлы не найдены в директории input")
                return
                
            logger.info(f"Найдено файлов для обработки: {len(word_files)}")
            
            for doc_path in word_files:
                try:
                    self.convert_single_file(doc_path)
                except Exception as e:
                    logger.error(f"Не удалось обработать файл {doc_path.name}: {str(e)}")
                    continue
                    
            logger.info("Обработка всех файлов завершена")
            
        except Exception as e:
            logger.error(f"Ошибка при пакетной обработке: {str(e)}", exc_info=True)
            raise

if __name__ == "__main__":
    try:
        converter = BatchConverter()
        converter.process_all_files()
    except Exception as e:
        logger.error(f"Программа завершилась с ошибкой: {str(e)}")
        raise